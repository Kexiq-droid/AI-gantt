from __future__ import annotations

import re
from datetime import date, datetime, timedelta
from io import BytesIO
from pathlib import Path
from typing import Any
from urllib.parse import quote

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from backend.app.seed_data import compute_schedule

HEADERS = [
    "код",
    "задача",
    "описание",
    "исполнитель",
    "длительность",
    "% выполнения",
    "дата начала",
    "дата конца",
    "предшественники",
    "родитель",
]
HEADER_ALIASES = {
    "код": ("код", "code"),
    "задача": ("задача", "title", "название"),
    "описание": ("описание", "description"),
    "исполнитель": ("исполнитель", "assignee"),
    "длительность": ("длительность", "duration", "duration_days"),
    "% выполнения": (
        "% выполнения",
        "прогресс",
        "progress",
        "progress_pct",
        "%",
        "выполнение",
    ),
    "дата начала": (
        "дата начала",
        "начало",
        "start",
        "start_date",
        "start date",
        "дата старта",
    ),
    "дата конца": (
        "дата конца",
        "конец",
        "окончание",
        "end",
        "end_date",
        "finish",
        "finish_date",
        "дата окончания",
    ),
    "предшественники": ("предшественники", "predecessors"),
    "родитель": ("родитель", "parent"),
}

LOGO_PATH = Path(__file__).resolve().parent.parent / "assets" / "bioplan_logo.png"

# Brand colors (match frontend theme)
ACCENT = "0F766E"
PHASE_BG = "E6F7F4"
PHASE_FG = "134E4A"
HEADER_FG = "FFFFFF"
MUTED = "5C6B66"
TEXT = "1C2422"
BORDER = "DDD4C8"

COL_WIDTHS = {
    "A": 12,   # код
    "B": 38,   # задача
    "C": 48,   # описание
    "D": 16,   # исполнитель
    "E": 13,   # длительность
    "F": 12,   # % выполнения
    "G": 14,   # дата начала
    "H": 14,   # дата конца
    "I": 20,   # предшественники
    "J": 12,   # родитель
}

HEADER_ROW = 5  # 1-based: logo/title block above
N_COLS = len(HEADERS)


def task_end_date(start: date, duration_days: int) -> date:
    return start + timedelta(days=max(int(duration_days or 0), 0))


def export_filename(plan: dict[str, Any], when: date | None = None) -> str:
    """Safe download name: BioPlan_<title>_<YYYY-MM-DD>.xlsx"""
    when = when or date.today()
    title = str(plan.get("title") or "plan").strip() or "plan"
    title = re.sub(r'[\\/:*?"<>|]+', "_", title)
    title = re.sub(r"\s+", "_", title).strip("._")[:60] or "plan"
    return f"BioPlan_{title}_{when.isoformat()}.xlsx"


def content_disposition(filename: str) -> str:
    ascii_name = filename.encode("ascii", "ignore").decode("ascii") or "BioPlan.xlsx"
    return f"attachment; filename=\"{ascii_name}\"; filename*=UTF-8''{quote(filename)}"


def _parse_excel_date(value: Any) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if not text:
        return None
    # ISO
    try:
        return date.fromisoformat(text[:10])
    except ValueError:
        pass
    # DD.MM.YYYY / DD/MM/YYYY
    for sep in (".", "/", "-"):
        parts = text.split(sep)
        if len(parts) == 3 and all(p.isdigit() for p in parts):
            a, b, c = (int(parts[0]), int(parts[1]), int(parts[2]))
            if c > 31:  # DD.MM.YYYY
                try:
                    return date(c, b, a)
                except ValueError:
                    return None
            if a > 31:  # YYYY-MM-DD already tried
                try:
                    return date(a, b, c)
                except ValueError:
                    return None
    return None


def export_plan_xlsx(plan: dict[str, Any]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "План"

    thin = Side(style="thin", color=BORDER)
    grid = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor=ACCENT)
    header_font = Font(name="Calibri", bold=True, color=HEADER_FG, size=11)
    phase_fill = PatternFill("solid", fgColor=PHASE_BG)
    phase_font = Font(name="Calibri", bold=True, color=PHASE_FG, size=11)
    body_font = Font(name="Calibri", color=TEXT, size=11)
    title_font = Font(name="Calibri", bold=True, color=ACCENT, size=16)
    meta_font = Font(name="Calibri", color=MUTED, size=10)
    last_col = get_column_letter(N_COLS)

    # —— Brand header ——
    ws.merge_cells(f"C1:{last_col}1")
    ws["C1"] = "R&D планирование"
    ws["C1"].font = title_font
    ws["C1"].alignment = Alignment(vertical="center")

    ws.merge_cells(f"A2:{last_col}2")
    ws["A2"] = f"План: {plan.get('title') or 'Без названия'}"
    ws["A2"].font = Font(name="Calibri", bold=True, color=TEXT, size=12)

    ws.merge_cells(f"A3:{last_col}3")
    exported_at = datetime.now().strftime("%d.%m.%Y %H:%M")
    ws["A3"] = f"Экспорт: {exported_at}"
    ws["A3"].font = meta_font

    # Accent strip under brand block
    for col in range(1, N_COLS + 1):
        cell = ws.cell(row=4, column=col, value="")
        cell.fill = PatternFill("solid", fgColor=ACCENT)
    ws.row_dimensions[1].height = 36
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 16
    ws.row_dimensions[4].height = 6

    if LOGO_PATH.is_file():
        logo = XLImage(str(LOGO_PATH))
        logo.width = 150
        logo.height = 44
        ws.add_image(logo, "A1")
    else:
        ws.merge_cells("A1:B1")
        ws["A1"] = "BioPlan"
        ws["A1"].font = title_font
        ws["A1"].alignment = Alignment(vertical="center")

    # —— Table header ——
    for col_idx, name in enumerate(HEADERS, start=1):
        cell = ws.cell(row=HEADER_ROW, column=col_idx, value=name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = grid
    ws.row_dimensions[HEADER_ROW].height = 22

    # —— Data ——
    for i, t in enumerate(plan.get("tasks") or []):
        row = HEADER_ROW + 1 + i
        is_phase = not t.get("parent")
        start = _parse_excel_date(t.get("start_date"))
        duration = int(t.get("duration_days") or 1)
        end = _parse_excel_date(t.get("end_date"))
        if start and not end:
            end = task_end_date(start, duration)
        values = [
            t["code"],
            t["title"],
            t.get("description") or "",
            t.get("assignee") or "",
            duration,
            max(0, min(100, int(t.get("progress_pct") or 0))),
            start,
            end,
            ",".join(t.get("predecessors") or []),
            t.get("parent") or "",
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.border = grid
            cell.font = phase_font if is_phase else body_font
            if is_phase:
                cell.fill = phase_fill
            wrap = col_idx in (2, 3)
            indent = 0 if is_phase or col_idx != 2 else 1
            cell.alignment = Alignment(
                vertical="center",
                wrap_text=wrap,
                indent=indent,
                horizontal="center" if col_idx in (5, 6, 7, 8) else "general",
            )
            if col_idx in (7, 8) and isinstance(value, date):
                cell.number_format = "DD.MM.YYYY"
        desc = str(t.get("description") or "")
        if len(desc) > 60:
            ws.row_dimensions[row].height = 36
        elif is_phase:
            ws.row_dimensions[row].height = 20

    last_data_row = HEADER_ROW + len(plan.get("tasks") or [])
    if last_data_row < HEADER_ROW:
        last_data_row = HEADER_ROW

    for letter, width in COL_WIDTHS.items():
        ws.column_dimensions[letter].width = width

    ws.freeze_panes = f"A{HEADER_ROW + 1}"
    ws.auto_filter.ref = f"A{HEADER_ROW}:{last_col}{max(last_data_row, HEADER_ROW)}"
    ws.print_title_rows = f"{HEADER_ROW}:{HEADER_ROW}"

    note_row = last_data_row + 2
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=N_COLS)
    note = ws.cell(
        row=note_row,
        column=1,
        value="Подсказка: при импорте сохраните заголовки. "
        "Колонки «дата начала» / «дата конца» задают сроки; "
        "если их нет — даты считаются по длительности и предшественникам.",
    )
    note.font = Font(name="Calibri", italic=True, color=MUTED, size=9)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _find_header_row(rows: list[tuple[Any, ...]]) -> tuple[int, list[str]]:
    """Locate the table header among brand/meta rows. Returns (index, lowered headers)."""
    title_names = HEADER_ALIASES["задача"]
    for idx, row in enumerate(rows[:30]):
        cells = [str(c or "").strip().lower() for c in row]
        if any(h in cells for h in title_names):
            return idx, cells
    raise ValueError(
        "Не найдена строка заголовков (ожидается колонка «задача»; "
        "опционально: код, описание, исполнитель, длительность, предшественники, …)"
    )


def _col_index(header: list[str], *names: str) -> int:
    for n in names:
        if n in header:
            return header.index(n)
    raise ValueError(f"Нет колонки {names[0]}")


def _optional_col(header: list[str], *names: str) -> int | None:
    for n in names:
        if n in header:
            return header.index(n)
    return None


def _next_auto_code(used: set[str]) -> str:
    n = 1
    while f"T{n}" in used:
        n += 1
    return f"T{n}"


def _resolve_pred_refs(
    raw_refs: list[str],
    *,
    by_code: dict[str, str],
    by_title: dict[str, str],
    by_index: dict[int, str],
) -> list[str]:
    """Map predecessor tokens to task codes (by code, title, or 1-based row index)."""
    out: list[str] = []
    for ref in raw_refs:
        if ref in by_code:
            out.append(ref)
            continue
        title_key = ref.casefold()
        if title_key in by_title:
            out.append(by_title[title_key])
            continue
        if ref.isdigit():
            code = by_index.get(int(ref))
            if code:
                out.append(code)
                continue
        out.append(ref)  # keep as-is; validate_plan will report unknown codes
    return out


def import_plan_xlsx(content: bytes, plan_start: date | None = None) -> dict[str, Any]:
    wb = load_workbook(BytesIO(content), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Пустой Excel")

    header_idx, header = _find_header_row(rows)

    i_code = _optional_col(header, *HEADER_ALIASES["код"])
    i_title = _col_index(header, *HEADER_ALIASES["задача"])
    i_desc = _optional_col(header, *HEADER_ALIASES["описание"])
    i_assignee = _optional_col(header, *HEADER_ALIASES["исполнитель"])
    i_dur = _optional_col(header, *HEADER_ALIASES["длительность"])
    i_progress = _optional_col(header, *HEADER_ALIASES["% выполнения"])
    i_start = _optional_col(header, *HEADER_ALIASES["дата начала"])
    i_end = _optional_col(header, *HEADER_ALIASES["дата конца"])
    i_pred = _optional_col(header, *HEADER_ALIASES["предшественники"])
    i_parent = _optional_col(header, *HEADER_ALIASES["родитель"])

    # Prefer title from brand header if present
    plan_title = "Импортированный план"
    for row in rows[:header_idx]:
        for cell in row:
            text = str(cell or "").strip()
            if text.lower().startswith("план:"):
                plan_title = text.split(":", 1)[1].strip() or plan_title
                break

    def cell(row: tuple[Any, ...], index: int | None) -> Any:
        if index is None:
            return None
        return row[index] if index < len(row) else None

    used_codes: set[str] = set()
    tasks: list[dict[str, Any]] = []
    for idx, row in enumerate(rows[header_idx + 1 :], start=header_idx + 2):
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue

        title_raw = str(cell(row, i_title) or "").strip()
        code = str(cell(row, i_code) or "").strip() if i_code is not None else ""

        # Skip footer hints / non-data rows (garbage in code column)
        if code and " " in code and not re.match(r"^[PT]\d", code, re.I):
            continue
        # Need at least a title or an explicit code
        if not title_raw and not code:
            continue

        if not code:
            code = _next_auto_code(used_codes)
        if code in used_codes:
            raise ValueError(f"Дублирующийся код задачи: {code}")
        used_codes.add(code)

        preds_raw = str(cell(row, i_pred) or "").strip()
        preds = [p.strip() for p in preds_raw.replace(";", ",").split(",") if p.strip()]
        parent = str(cell(row, i_parent) or "").strip() or None
        if not title_raw:
            title_raw = code

        start = _parse_excel_date(cell(row, i_start))
        end = _parse_excel_date(cell(row, i_end))
        dur_raw = cell(row, i_dur)
        try:
            duration = int(dur_raw) if dur_raw not in (None, "") else None
        except (TypeError, ValueError):
            duration = None

        # Prefer explicit dates from Excel when present
        if start and end and (duration is None or duration <= 0):
            duration = max((end - start).days, 1)
        elif start and end and duration and duration > 0:
            # Keep duration; end in UI is derived as start+duration
            pass
        elif start and duration is None and end is None:
            duration = 1
        elif end and duration and duration > 0 and start is None:
            start = end - timedelta(days=duration)
        elif duration is None:
            duration = 1

        duration = max(int(duration or 1), 1)

        progress = 0
        prog_raw = cell(row, i_progress)
        if prog_raw not in (None, ""):
            try:
                progress = int(float(str(prog_raw).replace("%", "").strip()))
            except (TypeError, ValueError):
                progress = 0
            progress = max(0, min(100, progress))

        tasks.append(
            {
                "code": code,
                "title": title_raw.strip(),
                "description": str(cell(row, i_desc) or "").strip(),
                "assignee": str(cell(row, i_assignee) or "").strip(),
                "duration_days": duration,
                "progress_pct": progress,
                "predecessors": preds,
                "parent": parent,
                "sort_order": idx * 10,
                "last_changed_by": "user",
                "_explicit_start": start,
            }
        )

    if not tasks:
        raise ValueError("В файле нет задач для импорта")

    # Resolve predecessors by code / title / 1-based row index among imported tasks
    by_code = {t["code"]: t["code"] for t in tasks}
    by_title: dict[str, str] = {}
    for t in tasks:
        key = str(t["title"]).casefold()
        by_title.setdefault(key, t["code"])
    by_index = {i + 1: t["code"] for i, t in enumerate(tasks)}
    for t in tasks:
        t["predecessors"] = _resolve_pred_refs(
            t["predecessors"],
            by_code=by_code,
            by_title=by_title,
            by_index=by_index,
        )
        if t["parent"]:
            parent_ref = t["parent"]
            if parent_ref not in by_code:
                resolved = by_title.get(parent_ref.casefold()) or (
                    by_index.get(int(parent_ref)) if parent_ref.isdigit() else None
                )
                if resolved:
                    t["parent"] = resolved

    fallback_start = plan_start or date.today()
    explicit_starts = [t["_explicit_start"] for t in tasks if t.get("_explicit_start")]
    if explicit_starts:
        fallback_start = min(explicit_starts)

    # Schedule only tasks without explicit start; keep Excel dates for the rest
    need_schedule = [t for t in tasks if not t.get("_explicit_start")]
    computed: dict[str, date] = {}
    if need_schedule:
        # Provide full task list so predecessors resolve, but overwrite only missing
        computed = compute_schedule(tasks, fallback_start)

    for t in tasks:
        start = t.pop("_explicit_start", None) or computed.get(t["code"]) or fallback_start
        t["start_date"] = start.isoformat()

    return {
        "title": plan_title,
        "start_date": fallback_start.isoformat(),
        "tasks": tasks,
    }
