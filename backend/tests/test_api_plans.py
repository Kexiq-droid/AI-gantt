"""HTTP plans: current, patch, shift, import/export, undo/redo, reset-seed."""

from datetime import date
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

EXAMPLE_XLSX = Path(__file__).resolve().parents[2] / "examples" / "plan_vax_b_demo.xlsx"


def test_get_current_plan(login_pm):
    r = login_pm.get("/api/plans/current")
    assert r.status_code == 200
    body = r.json()
    assert body["title"]
    assert len(body["tasks"]) > 0


def test_patch_task_title_assignee(login_pm):
    plan = login_pm.get("/api/plans/current").json()
    task = next(t for t in plan["tasks"] if t.get("parent_id") is not None)
    r = login_pm.patch(
        f"/api/plans/tasks/{task['id']}",
        json={"title": "Patched title", "assignee": "Тестов"},
    )
    assert r.status_code == 200
    updated = next(t for t in r.json()["tasks"] if t["id"] == task["id"])
    assert updated["title"] == "Patched title"
    assert updated["assignee"] == "Тестов"


def test_shift_tasks(login_pm):
    plan = login_pm.get("/api/plans/current").json()
    task = next(t for t in plan["tasks"] if t["code"] == "T2.1")
    old = task["start_date"]
    r = login_pm.post(
        "/api/plans/tasks/shift",
        json={"task_ids": [task["id"]], "days": 5},
    )
    assert r.status_code == 200
    new = next(t for t in r.json()["tasks"] if t["id"] == task["id"])
    assert new["start_date"] != old


def test_import_demo_xlsx(login_pm):
    content = EXAMPLE_XLSX.read_bytes()
    r = login_pm.post(
        "/api/plans/current/import",
        files={
            "file": (
                "plan_vax_b_demo.xlsx",
                content,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert r.status_code == 200, r.text
    assert len(r.json()["tasks"]) > 0


def test_export_has_date_columns(login_pm):
    r = login_pm.get("/api/plans/current/export")
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers.get("content-type", "")
    wb = load_workbook(BytesIO(r.content))
    ws = wb.active
    headers = [str(c.value or "").strip().lower() for c in next(ws.iter_rows(min_row=5, max_row=5))]
    assert "дата начала" in headers
    assert "дата конца" in headers


def test_undo_redo(login_pm):
    plan = login_pm.get("/api/plans/current").json()
    task = next(t for t in plan["tasks"] if t.get("parent_id") is not None)
    original = task["title"]
    code = task["code"]
    login_pm.patch(f"/api/plans/tasks/{task['id']}", json={"title": "UndoMe"})
    r = login_pm.post("/api/plans/current/undo")
    assert r.status_code == 200
    restored = next(t for t in r.json()["tasks"] if t["code"] == code)
    assert restored["title"] == original
    r2 = login_pm.post("/api/plans/current/redo")
    assert r2.status_code == 200
    redone = next(t for t in r2.json()["tasks"] if t["code"] == code)
    assert redone["title"] == "UndoMe"


def test_reset_seed(login_pm):
    login_pm.patch(
        f"/api/plans/tasks/{login_pm.get('/api/plans/current').json()['tasks'][0]['id']}",
        json={"title": "Dirty"},
    )
    r = login_pm.post("/api/plans/current/reset-seed")
    assert r.status_code == 200
    body = r.json()
    assert body["start_date"] == date.today().isoformat()
    assert len(body["tasks"]) > 0
    assert any(t["code"] == "P1" for t in body["tasks"])


def test_create_task(login_pm):
    plan = login_pm.get("/api/plans/current").json()
    parent = next(t for t in plan["tasks"] if t["code"] == "P2")
    after = next(t for t in plan["tasks"] if t["code"] == "T2.1")
    r = login_pm.post(
        "/api/plans/tasks",
        json={
            "title": "Новая задача UI",
            "parent_id": parent["id"],
            "after_task_id": after["id"],
            "duration_days": 3,
            "assignee": "Тестов",
            "start_date": "2026-08-01",
        },
    )
    assert r.status_code == 200, r.text
    created = next(t for t in r.json()["tasks"] if t["title"] == "Новая задача UI")
    assert created["parent_id"] == parent["id"]
    assert created["code"].startswith("P2.")
    assert created["assignee"] == "Тестов"
    assert created["start_date"] == "2026-08-01"
    assert created["duration_days"] == 3
    assert created["end_date"] == "2026-08-04"


def test_delete_leaf_and_phase_forbidden(login_pm):
    plan = login_pm.get("/api/plans/current").json()
    leaf = next(t for t in plan["tasks"] if t["code"] == "T2.1")
    phase = next(t for t in plan["tasks"] if t["code"] == "P2")
    bad = login_pm.delete(f"/api/plans/tasks/{phase['id']}")
    assert bad.status_code == 400
    ok = login_pm.delete(f"/api/plans/tasks/{leaf['id']}")
    assert ok.status_code == 200
    codes = {t["code"] for t in ok.json()["tasks"]}
    assert "T2.1" not in codes


def test_reorder_siblings(login_pm):
    plan = login_pm.get("/api/plans/current").json()
    t21 = next(t for t in plan["tasks"] if t["code"] == "T2.1")
    t22 = next(t for t in plan["tasks"] if t["code"] == "T2.2")
    # Move T2.1 after T2.2
    r = login_pm.post(
        "/api/plans/tasks/reorder",
        json={"task_id": t21["id"], "after_task_id": t22["id"]},
    )
    assert r.status_code == 200, r.text
    sibs = sorted(
        [t for t in r.json()["tasks"] if t["parent_id"] == t21["parent_id"]],
        key=lambda t: t["sort_order"],
    )
    codes = [t["code"] for t in sibs]
    assert codes.index("T2.2") < codes.index("T2.1")
