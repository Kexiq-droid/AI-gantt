from datetime import date
from io import BytesIO

from openpyxl import Workbook, load_workbook

from backend.app.services.excel_io import (
    export_filename,
    export_plan_xlsx,
    import_plan_xlsx,
)
from backend.app.services.validate import validate_plan_dict


def _sample_plan():
    return {
        "title": "Demo R&D",
        "start_date": "2026-03-02",
        "tasks": [
            {
                "code": "P1",
                "parent": None,
                "title": "Фаза",
                "description": "d",
                "assignee": "",
                "duration_days": 5,
                "start_date": "2026-03-02",
                "sort_order": 1,
                "predecessors": [],
            },
            {
                "code": "T1.1",
                "parent": "P1",
                "title": "Задача",
                "description": "desc",
                "assignee": "Иванов",
                "duration_days": 3,
                "start_date": "2026-03-02",
                "sort_order": 2,
                "predecessors": [],
            },
            {
                "code": "T1.2",
                "parent": "P1",
                "title": "После",
                "description": "",
                "assignee": "Петрова",
                "duration_days": 2,
                "start_date": "2026-03-05",
                "sort_order": 3,
                "predecessors": ["T1.1"],
            },
        ],
    }


def test_excel_roundtrip():
    plan = _sample_plan()
    raw = export_plan_xlsx(plan)
    imported = import_plan_xlsx(raw, plan_start=date(2026, 3, 2))
    assert not validate_plan_dict(imported)
    codes = {t["code"] for t in imported["tasks"]}
    assert codes == {"P1", "T1.1", "T1.2"}
    t12 = next(t for t in imported["tasks"] if t["code"] == "T1.2")
    assert "T1.1" in t12["predecessors"]
    assert t12["parent"] == "P1"
    assert imported["title"] == "Demo R&D"
    p1 = next(t for t in imported["tasks"] if t["code"] == "P1")
    assert p1["title"] == "Фаза"
    assert p1["parent"] is None
    assert p1["start_date"] == "2026-03-02"
    assert t12["start_date"] == "2026-03-05"


def test_export_includes_start_end_columns():
    raw = export_plan_xlsx(_sample_plan())
    wb = load_workbook(BytesIO(raw), data_only=True)
    ws = wb.active
    headers = [str(c.value or "").strip().lower() for c in next(ws.iter_rows(min_row=5, max_row=5))]
    assert "дата начала" in headers
    assert "дата конца" in headers
    # T1.1: start 02.03.2026, duration 3 → end 05.03.2026
    i_start = headers.index("дата начала")
    i_end = headers.index("дата конца")
    row = [c.value for c in next(ws.iter_rows(min_row=7, max_row=7))]  # T1.1
    start_v = row[i_start].date() if hasattr(row[i_start], "date") else row[i_start]
    end_v = row[i_end].date() if hasattr(row[i_end], "date") else row[i_end]
    assert start_v == date(2026, 3, 2)
    assert end_v == date(2026, 3, 5)


def test_import_uses_explicit_dates():
    wb = Workbook()
    ws = wb.active
    ws.append(
        [
            "код",
            "задача",
            "описание",
            "исполнитель",
            "длительность",
            "дата начала",
            "дата конца",
            "предшественники",
            "родитель",
        ]
    )
    ws.append(["P1", "Фаза", "", "", 10, date(2026, 4, 1), date(2026, 4, 11), "", ""])
    ws.append(["T1.1", "Задача", "", "Иванов", 4, date(2026, 4, 3), date(2026, 4, 7), "", "P1"])
    buf = BytesIO()
    wb.save(buf)
    imported = import_plan_xlsx(buf.getvalue(), plan_start=date(2026, 1, 1))
    assert imported["start_date"] == "2026-04-01"
    t11 = next(t for t in imported["tasks"] if t["code"] == "T1.1")
    assert t11["start_date"] == "2026-04-03"
    assert t11["duration_days"] == 4


def test_import_derives_duration_from_dates():
    wb = Workbook()
    ws = wb.active
    ws.append(
        [
            "код",
            "задача",
            "описание",
            "исполнитель",
            "длительность",
            "дата начала",
            "дата конца",
            "предшественники",
            "родитель",
        ]
    )
    ws.append(["T1", "Alone", "", "", "", "2026-05-10", "2026-05-15", "", ""])
    buf = BytesIO()
    wb.save(buf)
    imported = import_plan_xlsx(buf.getvalue())
    t = imported["tasks"][0]
    assert t["start_date"] == "2026-05-10"
    assert t["duration_days"] == 5


def test_excel_legacy_plain_header_still_imports():
    """Old draft exports (header on row 1, no date cols) must keep working."""
    wb = Workbook()
    ws = wb.active
    ws.append(["код", "задача", "описание", "исполнитель", "длительность", "предшественники", "родитель"])
    ws.append(["P1", "Фаза", "", "", 5, "", ""])
    ws.append(["T1.1", "Задача", "desc", "Иванов", 3, "", "P1"])
    buf = BytesIO()
    wb.save(buf)
    imported = import_plan_xlsx(buf.getvalue(), plan_start=date(2026, 3, 2))
    assert {t["code"] for t in imported["tasks"]} == {"P1", "T1.1"}
    assert imported["tasks"][0]["start_date"] == "2026-03-02"


def test_export_filename_sanitizes():
    name = export_filename({"title": 'План: тест/А"B'}, when=date(2026, 7, 28))
    assert name.startswith("BioPlan_")
    assert name.endswith("_2026-07-28.xlsx")
    assert "/" not in name
    assert '"' not in name


def test_import_tz_five_columns_only():
    """Assignment Excel: задача, описание, исполнитель, длительность, предшественники."""
    wb = Workbook()
    ws = wb.active
    ws.append(["задача", "описание", "исполнитель", "длительность", "предшественники"])
    ws.append(["Анализ рынка", "обзор", "Петрова", 5, ""])
    ws.append(["Синтез", "лаб", "Иванов", 3, "Анализ рынка"])
    buf = BytesIO()
    wb.save(buf)
    imported = import_plan_xlsx(buf.getvalue(), plan_start=date(2026, 3, 2))
    assert not validate_plan_dict(imported)
    assert [t["code"] for t in imported["tasks"]] == ["T1", "T2"]
    assert imported["tasks"][0]["assignee"] == "Петрова"
    assert imported["tasks"][1]["predecessors"] == ["T1"]
    assert imported["tasks"][1]["start_date"] > imported["tasks"][0]["start_date"]


def test_import_without_parent_and_pred_columns():
    wb = Workbook()
    ws = wb.active
    ws.append(["код", "задача", "описание", "исполнитель", "длительность"])
    ws.append(["T1", "Alone", "d", "Иванов", 4])
    buf = BytesIO()
    wb.save(buf)
    imported = import_plan_xlsx(buf.getvalue(), plan_start=date(2026, 3, 2))
    assert not validate_plan_dict(imported)
    assert imported["tasks"][0]["code"] == "T1"
    assert imported["tasks"][0]["predecessors"] == []
    assert imported["tasks"][0]["parent"] is None


def test_import_examples_sample_file():
    from pathlib import Path

    raw = (Path(__file__).resolve().parents[2] / "examples" / "plan_biokad_demo.xlsx").read_bytes()
    imported = import_plan_xlsx(raw)
    assert not validate_plan_dict(imported)
    assert len(imported["tasks"]) >= 10


def test_import_pred_by_row_index():
    wb = Workbook()
    ws = wb.active
    ws.append(["задача", "описание", "исполнитель", "длительность", "предшественники"])
    ws.append(["A", "", "", 2, ""])
    ws.append(["B", "", "", 2, "1"])
    buf = BytesIO()
    wb.save(buf)
    imported = import_plan_xlsx(buf.getvalue(), plan_start=date(2026, 1, 1))
    assert imported["tasks"][1]["predecessors"] == ["T1"]
    assert not validate_plan_dict(imported)
